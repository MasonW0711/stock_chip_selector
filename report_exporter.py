# ==========================================
# report_exporter.py - Excel 報表匯出
# ==========================================

import io
import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, Any

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# 報表標題列的填色（深藍）
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)

# 高分列的填色（淺黃）
HIGHLIGHT_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')


def export_to_excel(
    result_df: pd.DataFrame,
    broker_detail_df: pd.DataFrame,
    holder_history_df: pd.DataFrame,
    params: Dict[str, Any]
) -> bytes:
    """
    產生完整 Excel 報表，包含四個工作表。

    工作表：
    1. 選股結果總表  - 所有符合條件的股票，依分數排序
    2. 分點買超明細  - 每筆 streak 的每日買超記錄
    3. 集保戶數變化  - 近 N 週集保戶數歷史
    4. 參數設定紀錄  - 本次篩選使用的參數

    回傳:
        Excel 檔案的 bytes 物件（可直接傳給 st.download_button）
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        _write_result_sheet(writer, result_df)
        _write_broker_detail_sheet(writer, broker_detail_df)
        _write_holder_sheet(writer, holder_history_df)
        _write_params_sheet(writer, params)

    return output.getvalue()


def _write_result_sheet(writer: pd.ExcelWriter, result_df: pd.DataFrame) -> None:
    """寫入選股結果總表（Sheet 1）"""
    sheet_name = '選股結果總表'

    if result_df.empty:
        pd.DataFrame({'提示': ['無符合條件的股票']}).to_excel(
            writer, sheet_name=sheet_name, index=False
        )
        return

    # 欄位中文對照
    col_mapping = {
        'stock_id': '股票代號',
        'stock_name': '股票名稱',
        'branch': '券商分點',
        'streak_days': '連續買超天數',
        'streak_period': '連續買超期間',
        'total_net_buy': '總買超張數',
        'avg_cost': '平均買進成本',
        'latest_close': '最新收盤價',
        'price_deviation_pct': '現價偏離率(%)',
        'early_holder': '四週前集保戶數',
        'latest_holder': '最新集保戶數',
        'holder_change_rate': '集保戶數變化率(%)',
        'is_increasing': '是否逐日增加',
        'trend_type': '買超趨勢',
        'score': '選股分數',
        'label': '評級',
        'price_label': '成本評語',
    }

    # 選取並重新命名欄位
    available = [c for c in col_mapping if c in result_df.columns]
    display_df = result_df[available].rename(columns=col_mapping)

    # 依選股分數由高到低排序
    if '選股分數' in display_df.columns:
        display_df = display_df.sort_values('選股分數', ascending=False)

    display_df.to_excel(writer, sheet_name=sheet_name, index=False)

    ws = writer.sheets[sheet_name]
    _apply_header_style(ws)
    _auto_column_width(ws)

    # 高分列（>=85分）標記淺黃色
    if '選股分數' in display_df.columns:
        score_col_idx = list(display_df.columns).index('選股分數') + 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            score_cell = ws.cell(row=row_idx, column=score_col_idx)
            try:
                if score_cell.value and int(score_cell.value) >= 85:
                    for cell in row:
                        cell.fill = HIGHLIGHT_FILL
            except (TypeError, ValueError):
                pass


def _write_broker_detail_sheet(writer: pd.ExcelWriter, broker_detail_df: pd.DataFrame) -> None:
    """寫入分點買超明細（Sheet 2）"""
    sheet_name = '分點買超明細'

    if broker_detail_df.empty:
        pd.DataFrame({'提示': ['無買超明細資料']}).to_excel(
            writer, sheet_name=sheet_name, index=False
        )
        return

    col_mapping = {
        'date': '日期',
        'stock_id': '股票代號',
        'stock_name': '股票名稱',
        'broker': '券商名稱',
        'branch': '分點名稱',
        'buy_volume': '買進張數',
        'sell_volume': '賣出張數',
        'net_buy': '買超張數',
        'buy_avg_price': '申報買進均價',
        'effective_buy_price': '計算用買進均價',
    }

    available = [c for c in col_mapping if c in broker_detail_df.columns]
    display_df = broker_detail_df[available].rename(columns=col_mapping)

    # 依股票代號與日期排序
    sort_cols = [c for c in ['股票代號', '分點名稱', '日期'] if c in display_df.columns]
    if sort_cols:
        display_df = display_df.sort_values(sort_cols)

    display_df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    _apply_header_style(ws)
    _auto_column_width(ws)


def _write_holder_sheet(writer: pd.ExcelWriter, holder_history_df: pd.DataFrame) -> None:
    """寫入集保戶數變化（Sheet 3）"""
    sheet_name = '集保戶數變化'

    if holder_history_df.empty:
        pd.DataFrame({'提示': ['無集保戶數資料']}).to_excel(
            writer, sheet_name=sheet_name, index=False
        )
        return

    col_mapping = {
        'week_date': '週別日期',
        'stock_id': '股票代號',
        'stock_name': '股票名稱',
        'holder_count': '集保戶數',
    }

    available = [c for c in col_mapping if c in holder_history_df.columns]
    display_df = holder_history_df[available].rename(columns=col_mapping)

    sort_cols = [c for c in ['股票代號', '週別日期'] if c in display_df.columns]
    if sort_cols:
        display_df = display_df.sort_values(sort_cols)

    display_df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    _apply_header_style(ws)
    _auto_column_width(ws)


def _write_params_sheet(writer: pd.ExcelWriter, params: Dict[str, Any]) -> None:
    """寫入本次篩選參數設定紀錄（Sheet 4）"""
    data_source = params.get('data_source', 'csv')
    data_source_label = 'FinMind sponsor API 自動下載' if data_source == 'finmind' else '手動 CSV 上傳'
    sheet_name = '參數設定紀錄'

    market_label = {'all': '全部', 'listed': '上市（TWSE）', 'otc': '上櫃（OTC）'}
    trend_label = '嚴格模式（買超張數逐日增加）' if params.get('strict_trend_mode') else '寬鬆模式（每天有買超即可）'

    params_data = {
        '參數名稱': [
            '篩選執行時間',
            '資料來源模式',
            '最小連續買超天數',
            '現價高於主力成本上限 (%)',
            '集保戶數觀察週數',
            '集保戶數最小下降比例 (%)',
            '買超趨勢模式',
            '最低成交量 (張)',
            '股票市場範圍',
        ],
        '設定值': [
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            data_source_label,
            params.get('min_consecutive_days', 3),
            params.get('max_price_deviation_pct', 5.0),
            params.get('holder_observation_weeks', 4),
            params.get('min_holder_decrease_pct', 0.0),
            trend_label,
            params.get('min_volume', 100),
            market_label.get(params.get('market_scope', 'all'), '全部'),
        ],
        '說明': [
            '自動記錄，供回測比對',
            '本次執行使用手動 CSV 或 FinMind sponsor API',
            '同一分點連續正買超至少幾天',
            '現價超過此比例則淘汰',
            '最新集保戶數 vs. N週前的比較週距',
            '0=只要有下降；5=下降須達5%以上',
            '嚴格=買超張數必須逐日遞增',
            '日平均成交量低於此值則過濾',
            '上市=TWSE；上櫃=OTC；全部=兩者皆納入',
        ]
    }

    if data_source == 'finmind':
        params_data['參數名稱'].extend([
            '原始輸入股票清單',
            '自動下載股票清單',
            '自動下載開始日期',
            '自動下載結束日期',
            '集保補抓起始日期',
        ])
        params_data['設定值'].extend([
            ', '.join(params.get('requested_stock_ids', params.get('broker_stock_ids', []))),
            ', '.join(params.get('broker_stock_ids', [])),
            params.get('broker_start_date', ''),
            params.get('broker_end_date', ''),
            params.get('holder_start_date', ''),
        ])
        params_data['說明'].extend([
            '使用者在自動下載模式輸入的原始股票代號清單',
            '本次向 FinMind 查詢的股票代號',
            '股價與券商分點資料的下載起點',
            '股價、集保、券商分點資料的下載終點',
            '為滿足觀察週數而自動往前延伸的集保資料起點',
        ])

    pd.DataFrame(params_data).to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    _apply_header_style(ws)
    _auto_column_width(ws)


def _apply_header_style(ws) -> None:
    """套用標題列樣式（深藍底白字置中）"""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 20


def _auto_column_width(ws) -> None:
    """依內容自動調整欄寬（最寬不超過 50）"""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    # 中文字符佔 2 個字元寬度
                    val_str = str(cell.value)
                    length = sum(2 if ord(c) > 127 else 1 for c in val_str)
                    if length > max_length:
                        max_length = length
            except Exception:
                pass
        adjusted_width = min(max_length + 4, 50)
        ws.column_dimensions[col_letter].width = max(adjusted_width, 8)
